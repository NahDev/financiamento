from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook


def calcular_financiamento():
    # Obter os valores digitados pelo usuário
    nome_carro = entry_nome.get()
    valor_carro = float(entry_valor.get())
    taxa_juros = float(entry_taxa.get())
    parcelas = int(entry_parcelas.get())

    # Calcular o valor da parcela e os juros mensais
    taxa_juros_mensal = taxa_juros / 100 / 12
    valor_parcela = (valor_carro * taxa_juros_mensal) / (
        1 - (1 + taxa_juros_mensal) ** -parcelas
    )

    # Criar uma planilha Excel
    wb = Workbook()
    planilha = wb.active
    planilha.title = "Financiamento"
    planilha["A1"] = "Parcela"
    planilha["B1"] = "Valor Parcela"
    planilha["C1"] = "Juros"
    planilha["D1"] = "Amortização"
    planilha["E1"] = "Saldo Devedor"

    saldo_devedor = valor_carro

    # Preencher a planilha com os valores de cada parcela
    for i in range(1, parcelas + 1):
        juros = saldo_devedor * taxa_juros_mensal
        amortizacao = valor_parcela - juros
        saldo_devedor -= amortizacao

        planilha.cell(row=i + 1, column=1, value=i)
        planilha.cell(row=i + 1, column=2, value=valor_parcela)
        planilha.cell(row=i + 1, column=3, value=juros)
        planilha.cell(row=i + 1, column=4, value=amortizacao)
        planilha.cell(row=i + 1, column=5, value=saldo_devedor)

    # Salvar o arquivo Excel
    nome_arquivo = f"{nome_carro}.xlsx"
    wb.save(nome_arquivo)
    messagebox.showinfo(
        "Financiamento",
        f"O financiamento foi calculado e salvo no arquivo '{nome_arquivo}'.",
    )


# Criar a interface gráfica com o tkinter
root = Tk()
root.title("Simulação de Financiamento")
root.geometry("400x200")

label_nome = Label(root, text="Nome do Carro:")
label_nome.pack()
entry_nome = Entry(root)
entry_nome.pack()

label_valor = Label(root, text="Valor do Carro:")
label_valor.pack()
entry_valor = Entry(root)
entry_valor.pack()

label_taxa = Label(root, text="Taxa de Juros (%):")
label_taxa.pack()
entry_taxa = Entry(root)
entry_taxa.pack()

label_parcelas = Label(root, text="Número de Parcelas:")
label_parcelas.pack()
entry_parcelas = Entry(root)
entry_parcelas.pack()

button_calcular = Button(
    root, text="Calcular Financiamento", command=calcular_financiamento
)
button_calcular.pack()

root.mainloop()
